#!/usr/bin/env python3
"""Fix the updated tracker: proper filename, clean TOP 추천, update summaries."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = "/Users/sam/Desktop/brain/자소서/홍석진 취업/채용트래커/Job_Tracker_20250628_updated.xlsx"
DST = "/Users/sam/Desktop/brain/자소서/홍석진 취업/채용트래커/Job_Tracker_20260628.xlsx"

wb = openpyxl.load_workbook(SRC)

# ─── Fix 1: 전체 리스트 - remove 서치펌 from new section, re-sort ───
ws = wb["전체 리스트"]

# Find the new section
new_section_start = None
new_section_end = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v and isinstance(v, str) and v.startswith("▌🆕 06/28"):
        new_section_start = r
    elif new_section_start and v and isinstance(v, str) and v.startswith("📊"):
        new_section_end = r - 1
        break

if new_section_start and not new_section_end:
    new_section_end = ws.max_row

print(f"New section: rows {new_section_start}-{new_section_end}")

# Read new jobs
new_jobs = []
for r in range(new_section_start + 1, new_section_end + 1):
    규모 = ws.cell(row=r, column=2).value
    회사명 = ws.cell(row=r, column=3).value
    직무 = ws.cell(row=r, column=4).value
    구분 = ws.cell(row=r, column=5).value
    고용형태 = ws.cell(row=r, column=6).value
    지역 = ws.cell(row=r, column=7).value
    마감일 = ws.cell(row=r, column=8).value
    스코어 = ws.cell(row=r, column=9).value
    링크_셀 = ws.cell(row=r, column=10)
    지원상태 = ws.cell(row=r, column=11).value
    부합도 = ws.cell(row=r, column=12).value
    
    if 회사명:
        new_jobs.append({
            "규모": 규모,
            "회사명": 회사명,
            "직무": 직무,
            "구분": 구분,
            "고용형태": 고용형태,
            "지역": 지역,
            "마감일": 마감일,
            "스코어": 스코어,
            "링크_url": str(링크_셀.hyperlink.target) if 링크_셀.hyperlink else "",
            "링크_text": 링크_셀.value,
            "지원상태": 지원상태,
            "내 부합도": 부합도,
        })

print(f"Read {len(new_jobs)} new jobs")

# Sort: 대기업 > 외국계 > 중견 > 중소, exclude 서치펌 from top
규모_order = {"대기업": 0, "외국계": 1, "중견": 2, "서치펌": 3, "중소": 4}
new_jobs.sort(key=lambda j: 규모_order.get(j["규모"], 5))

# Delete old new section rows (from bottom up)
for r in range(new_section_end, new_section_start - 1, -1):
    ws.delete_rows(r)

# Re-insert section header
section_header = f"▌🆕 06/28 신규 발굴 ({len(new_jobs)}건) — 원티드+피플앤잡"
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

insert_at = new_section_start
ws.insert_rows(insert_at)
ws.cell(row=insert_at, column=1).value = section_header
ws.cell(row=insert_at, column=1).font = header_font
ws.cell(row=insert_at, column=1).fill = header_fill
for c in range(2, 13):
    ws.cell(row=insert_at, column=c).fill = header_fill

# Insert jobs
규모_count = {}
for i, job in enumerate(new_jobs, 1):
    row = insert_at + i
    ws.cell(row=row, column=1).value = i
    ws.cell(row=row, column=2).value = job["규모"]
    ws.cell(row=row, column=3).value = job["회사명"]
    ws.cell(row=row, column=4).value = job["직무"]
    ws.cell(row=row, column=5).value = job.get("구분", "경력")
    ws.cell(row=row, column=6).value = job.get("고용형태", "정규")
    ws.cell(row=row, column=7).value = job.get("지역", "서울")
    ws.cell(row=row, column=8).value = job.get("마감일", "")
    ws.cell(row=row, column=9).value = job.get("스코어", 0)
    
    link_cell = ws.cell(row=row, column=10)
    link_cell.value = "🔗 지원하기"
    if job.get("링크_url"):
        link_cell.hyperlink = job["링크_url"]
        link_cell.font = Font(color="0563C1", underline="single")
    
    ws.cell(row=row, column=11).value = job.get("지원상태", "")
    ws.cell(row=row, column=12).value = job.get("내 부합도", "")
    
    # Count
    규모_count[job["규모"]] = 규모_count.get(job["규모"], 0) + 1

# Update summary rows
# Find summary rows
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v and isinstance(v, str) and v.startswith("📊 총"):
        # Count all jobs
        대기업_c = 0; 외국계_c = 0; 서치펌_c = 0; 중견_c = 0; 중소_c = 0
        for r2 in range(2, ws.max_row + 1):
            v2 = ws.cell(row=r2, column=1).value
            if v2 and isinstance(v2, str) and v2.startswith("▌"):
                continue
            if v2 and isinstance(v2, str) and v2.startswith("📊"):
                continue
            if v2 and isinstance(v2, str) and v2.startswith("📌"):
                continue
            규모_v = ws.cell(row=r2, column=2).value
            if 규모_v == "대기업": 대기업_c += 1
            elif 규모_v == "외국계": 외국계_c += 1
            elif 규모_v == "서치펌": 서치펌_c += 1
            elif 규모_v == "중견": 중견_c += 1
            elif 규모_v == "중소": 중소_c += 1
        
        total = 대기업_c + 외국계_c + 서치펌_c + 중견_c + 중소_c
        ws.cell(row=r, column=1).value = f"📊 총 {total}건 | 대기업 {대기업_c}건 · 외국계 {외국계_c}건 · 서치펌 {서치펌_c}건 · 중견기업 {중견_c}건 · 중소기업 {중소_c}건"
        
        # Merge across
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        ws.cell(row=r, column=1).font = Font(bold=True, size=10, color="666666")
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    
    if v and isinstance(v, str) and v.startswith("📌 크론잡"):
        ws.cell(row=r, column=1).value = f"📌 크론잡 매주 월요일 자동 모니터링 · 2026-06-28 업데이트 · 규모별 정렬"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        ws.cell(row=r, column=1).font = Font(bold=False, size=9, color="999999")
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")


# ─── Fix 2: TOP 추천 - remove 서치펌 ───
ws2 = wb["TOP 추천"]
# Find and remove 서치펌 section from TOP 추천
rows_to_delete = []
for r in range(1, ws2.max_row + 1):
    규모_v = ws2.cell(row=r, column=2).value
    section_v = ws2.cell(row=r, column=1).value
    
    # Check if this is a 서치펌 entry or the 신규 section with 서치펌 jobs
    if section_v and isinstance(section_v, str) and "신규" in section_v:
        # Check the next rows for 서치펌
        next_rows_are_서치펌 = True
        for r2 in range(r+1, min(r+5, ws2.max_row+1)):
            v = ws2.cell(row=r2, column=2).value
            if v and v != "서치펌":
                next_rows_are_서치펌 = False
                break
        if next_rows_are_서치펌:
            # Delete this section
            end_row = r
            for r2 in range(r+1, ws2.max_row+1):
                if ws2.cell(row=r2, column=2).value:
                    end_row = r2
                else:
                    break
            for r_del in range(end_row, r-1, -1):
                ws2.delete_rows(r_del)
            break

# Add clean TOP 추천 신규 section
# Only add 대기업 and 외국계 new jobs
top_new = [j for j in new_jobs if j["규모"] in ("대기업", "외국계")][:15]
if top_new:
    insert_top = ws2.max_row + 1
    ws2.insert_rows(insert_top)
    ws2.cell(row=insert_top, column=1).value = f"▌🆕 06/28 신규 ({len(top_new)}건)"
    ws2.cell(row=insert_top, column=1).font = header_font
    ws2.cell(row=insert_top, column=1).fill = header_fill
    for c in range(2, 13):
        ws2.cell(row=insert_top, column=c).fill = header_fill
    
    for i, job in enumerate(top_new, 1):
        row = insert_top + i
        ws2.cell(row=row, column=1).value = i
        ws2.cell(row=row, column=2).value = job["규모"]
        ws2.cell(row=row, column=3).value = job["회사명"]
        ws2.cell(row=row, column=4).value = job["직무"]
        ws2.cell(row=row, column=5).value = job.get("구분", "경력")
        ws2.cell(row=row, column=6).value = job.get("고용형태", "정규")
        ws2.cell(row=row, column=7).value = job.get("지역", "서울")
        ws2.cell(row=row, column=8).value = job.get("마감일", "")
        ws2.cell(row=row, column=9).value = job.get("스코어", 0)
        
        link_cell = ws2.cell(row=row, column=10)
        link_cell.value = "🔗 지원하기"
        if job.get("링크_url"):
            link_cell.hyperlink = job["링크_url"]
            link_cell.font = Font(color="0563C1", underline="single")
        
        ws2.cell(row=row, column=11).value = job.get("지원상태", "")
        ws2.cell(row=row, column=12).value = job.get("내 부합도", "")

# Save
wb.save(DST)
print(f"Saved to: {DST}")
print(f"규모 분포: {규모_count}")
