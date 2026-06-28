#!/usr/bin/env python3
"""
원티드 API + peoplenjob + LinkedIn 채용공고 수집 후 엑셀 트래커 업데이트
"""
import json, time, re, sys, os, hashlib
from datetime import datetime
from urllib.parse import quote

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ─── Config ───────────────────────────────────────────────────
TRACKER_PATH = "/Users/sam/Desktop/brain/자소서/홍석진 취업/채용트래커/Job_Tracker_20260628.xlsx"
OUTPUT_PATH = "/Users/sam/Desktop/brain/자소서/홍석진 취업/채용트래커/Job_Tracker_20260628.xlsx"  # overwrite same file (saved as new below)
NEW_TRACKER_PATH = "/Users/sam/Desktop/brain/자소서/홍석진 취업/채용트래커/Job_Tracker_20250628_updated.xlsx"
GIT_REPO = "/Users/sam/Desktop/brain/자소서/홍석진 취업/채용트래커"

TODAY = datetime.now().strftime("%Y-%m-%d")
TODAY_SHORT = datetime.now().strftime("%m/%d")

# Headers
WANTED_HEADERS = {
    "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 Mobile/15E148",
    "Accept": "application/json",
}
PEOPLE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
}

# Keywords for Wanted API
WANTED_KEYWORDS = [
    "사업개발", "Business Development", "해외영업", "B2B 영업", "Commercial",
    "Global Alliance", "Market Intelligence", "Key Account", "Partnership",
    "무역", "Trading", "SCM", "공급망", "전략기획", "러시아", "BD",
    "화학영업", "소재영업", "에너지", "식품영업", "바이오영업"
]

# Keywords for peoplenjob
PEOPLE_KEYWORDS = [
    "해외영업", "사업개발", "Business Development", "Commercial", "B2B", "무역", "글로벌", "영업"
]

# Existing company names in tracker (to skip duplicates)
EXISTING_COMPANIES = set()
EXISTING_POSITIONS = set()

# ─── 대기업 / 외국계 / 중견 기업 리스트 ──────────────────────
대기업_리스트 = {
    "삼성", "현대", "LG", "SK", "롯데", "포스코", "POSCO", "한화", "GS", "KT", "한진",
    "CJ", "두산", "LS", "DL", "효성", "LX", "OCI", "네이버", "카카오", "넥슨", "NC소프트",
    "셀트리온", "삼양", "대상", "농심", "KT&G", "현대모비스", "현대차", "기아",
    "SK이노베이션", "SK에너지", "SK지오센트릭", "SK온", "SK하이닉스", "SK텔레콤",
    "SKC", "SK케미칼", "SK바이오사이언스", "SK바이오팜", "SK머티리얼즈", "SK실트론",
    "SK가스", "SK E&S", "SK어드밴스드", "SK디앤디", "SK쉴더스", "SK에코플랜트",
    "SK렌터카", "SK네트웍스", "SK플라즈마", "SK피유코어", "SK매직",
    "LG화학", "LG에너지솔루션", "LG전자", "LG디스플레이", "LG이노텍", "LG CNS",
    "롯데케미칼", "롯데정밀화학", "롯데에너지머티리얼즈", "롯데바이오로직스",
    "LX인터내셔널", "LX판토스", "LX하우시스", "LX MMA",
    "삼성전자", "삼성SDI", "삼성디스플레이", "삼성바이오로직스", "삼성바이오에피스",
    "삼성물산", "삼성중공업", "삼성엔지니어링", "제일기획", "호텔신라",
    "포스코인터내셔널", "포스코퓨처엠", "포스코DX", "포스코플로우",
    "한화솔루션", "한화에어로스페이스", "한화오션", "한화토탈에너지스", "한화첨단소재",
    "GS칼텍스", "GS리테일", "GS건설", "GS에너지",
    "CJ제일제당", "CJ대한통운", "CJ ENM", "CJ올리브영",
    "두산에너빌리티", "두산밥캣", "두산퓨얼셀",
    "HD한국조선해양", "HD현대중공업", "HD현대미포", "HD현대삼호", "HD현대오일뱅크",
    "현대건설", "현대엔지니어링", "현대글로비스", "현대트랜시스", "현대케피코",
    "현대오토에버", "현대위아", "현대스틸산업", "현대백화점",
    "신세계", "이마트", "아모레퍼시픽", "LG생활건강",
    "금호석유화학", "금호타이어", "이수화학", "이수그룹",
    "S-Oil", "에쓰오일", "S-OIL",
    "태광산업", "한국타이어", "대한항공", "아시아나항공",
    "미래에셋", "KB금융", "신한금융", "하나금융",
    "동원", "오뚜기", "SPC", "파리크라상",
    "코오롱", "OCI홀딩스", "유한킴벌리", "한국콜마",
    "한미약품", "대웅제약", "종근당", "GC녹십자", "유한양행",
    "현대중공업그룹", "한국조선해양",
    "현대제철", "현대비앤지스틸",
    "고려아연", "영풍", "LS전선", "LS일렉트릭", "LS MnM",
    # 대기업 계열사 키워드
    "SK피아이씨글로벌", "SK picglobal", "SK피아이씨",
    "현대모비스", "현대트랜시스", "현대케피코",
}

외국계_리스트 = {
    "Cargill", "ADM", "Bunge", "Louis Dreyfus", "COFCO", "Glencore",
    "Trafigura", "Vitol", "Mercuria", "Gunvor",
    "BP", "Shell", "ExxonMobil", "Chevron", "TotalEnergies", "Equinor",
    "BASF", "Dow", "DuPont", "Evonik", "DSM", "Firmenich", "Givaudan",
    "IFF", "Symrise", "Solvay", "Arkema", "LG", "Covestro", "SABIC",
    "LyondellBasell", "Brenntag", "Univar", "Azelis", "IMCD", "Biesterfeld",
    "Neste", "UPM", "Stora Enso", "Metsä",
    "Nestlé", "PepsiCo", "Unilever", "P&G", "Mondelez", "Mars",
    "Danone", "Kellogg", "General Mills", "Tyson", "JBS",
    "Tate & Lyle", "Ingredion", "Roquette", "Kerry", "Givaudan",
    "Siemens", "GE", "ABB", "Schneider Electric", "Emerson", "Honeywell",
    "3M", "Ecolab", "Nalco", "Suez", "Veolia",
    "Cisco", "Oracle", "SAP", "Microsoft", "IBM", "AWS", "Google",
    "Meta", "Apple", "Intel", "AMD", "Nvidia", "Qualcomm",
    "Johnson & Johnson", "Pfizer", "Roche", "Novartis", "Merck", "GSK",
    "Sanofi", "AstraZeneca", "Bayer", "Boehringer Ingelheim", "Eli Lilly",
    "Medtronic", "Boston Scientific", "Stryker", "Zimmer Biomet",
    "Caterpillar", "Komatsu", "Volvo", "Hitachi", "Atlas Copco",
    "Maersk", "DHL", "FedEx", "UPS", "DB Schenker", "Kuehne+Nagel",
    "L'Oréal", "Estée Lauder", "Shiseido", "Chanel", "LVMH",
    "Nike", "Adidas", "Puma", "VF Corporation",
    "Accenture", "McKinsey", "BCG", "Bain", "Deloitte", "PwC", "EY", "KPMG",
    "Citi", "HSBC", "Standard Chartered", "BNP Paribas", "Credit Agricole",
    "ING", "Deutsche Bank", "UBS", "JP Morgan", "Morgan Stanley", "Goldman Sachs",
    "Air Liquide", "Linde", "Praxair",
    # 외국계 한국법인 패턴
    "코카콜라", "한국쓰리엠", "한국바스프", "한국다우", "듀폰코리아",
    "한국에보닉", "로레알코리아", "유니레버코리아", "P&G Korea",
    "네슬레코리아", "한국화이자", "한국로슈", "사노피", "GSK Korea",
    "지멘스", "한국지멘스", "ABB Korea", "슈나이더일렉트릭코리아",
}

중견_리스트 = {
    "이수화학", "이수그룹", "이수앱지스", "이수페타시스",
    "OCI", "OCI홀딩스", "OCI파워",
    "한솔케미칼", "한솔그룹", "한솔테크닉스",
    "KPX케미칼", "KPX그룹",
    "미원상사", "미원스페셜티케미칼", "미원에스씨",
    "KCC", "KCC글라스", "KCC실리콘",
    "삼양그룹", "삼양사", "삼양홀딩스",
    "대상", "대상홀딩스",
    "동원시스템즈", "동원F&B",
    "한국콜마", "콜마비앤에이치",
    "코스맥스", "아모레퍼시픽",
    "셀트리온", "삼성바이오로직스",
    "태림포장", "신대양제지",
    "삼표시멘트", "아세아시멘트",
    "한국알콜", "영우디에스피", "덕산네오룩스",
    "KH바텍", "에이치엘비", "티에스아이",
    "동진쎄미켐", "원익", "피에스케이",
    "심텍", "대덕전자", "코리아써키트",
    "리노공업", "테크윙", "엘오티베큠",
}

# ─── Filter keywords ───────────────────────────────────────────
BLACKLIST_TITLES = [
    "개발자", "개발", "프로그래머", "프론트엔드", "백엔드", "DevOps", "데브옵스",
    "iOS", "Android", "AI", "ML", "데이터엔지니어", "Data Engineer",
    "연구원", "연구소", "R&D", "Researcher", "Scientist",
    "마케터", "마케팅", "Marketing", "Brand", "콘텐츠",
    "일본어", "日本語", "Japanese", "일본",
    "디자이너", "Designer", "UX", "UI",
    "HR", "인사팀", "채용담당", "recruiter",
    "재무", "회계", "Finance", "Accounting", "CPA",
    "법무", "Legal", "변호사",
]
BLACKLIST_SENIORITY = [
    "5년", "7년", "10년", "15년", "20년", "차장", "부장", "이사", "임원",
    "Manager", "Director", "Senior", "Sr.", "Principal", "Lead", "Head",
    "팀장", "실장", "본부장", "CEO", "CTO", "CFO", "VP", "Vice President",
    "경력 5", "경력 7", "경력 10",
]
BLACKLIST_EMPLOYMENT = [
    "계약직", "Contract", "인턴", "Intern", "아르바이트", "프리랜서", "Freelance",
    "파견직", "위촉직",
]
BLACKLIST_LOCATION = [
    "부산", "대구", "광주", "대전", "울산", "전주", "청주", "창원", "포항",
    "여수", "구미", "거제", "순천", "원주", "강릉", "제주", "익산",
    "충북", "충남", "전북", "전남", "경북", "경남", "강원", "제주도",
    "해외", "overseas", "베트남", "중국", "인도네시아",
]

# ─── Helper functions ──────────────────────────────────────────
def classify_company(name):
    """분류: 대기업 / 외국계 / 중견 / 중소 / 서치펌"""
    if not name:
        return "중소"
    name_clean = name.strip()
    
    # 서치펌 체크
    search_firm_keywords = ["헤딩", "커리어케어", "커리어", "HR", "인재", "스카우트", 
                            "파트너스", "서치", "로버트월터스", "Robert Walters", "Robert Half",
                            "맨파워", "Manpower", "아데코", "Adecco", "랜스태드", "Randstad",
                            "헤이즈", "Hays", "퍼솔", "Persol", "워크데이", "Workday",
                            "링크드", "Linked", "HRnet", "피플케어", "에이커", "Acre",
                            "HR컨설팅", "인사컨설팅", "채용대행", "헤드헌팅",
                            "잡코리아", "사람인", "원티드랩", "로켓펀치",
                            "리멤버", "Remember", "커리어빌더", "Careerbuilder",
                            "BRecruit", "비리크루트", "MJ Human", "유앤파트너즈",
                            "HRForce", "에이치알포스", "스톤브릿지", "Stonebridge",
                            "HR링크", "탑커리어", "TopCareer", "잡에이스", "JobAce",
                            "나우커리어", "NowCareer", "에코파트너스", "EchoPartners",
                            "HR파트너스", "파인드잡", "FindJob", "드림HR",
                            "JAC Recruitment", "JAC", "Michael Page", "PageGroup",
                            "아데코", "Adecco", "코리아HR", "KoreaHR",
                            "유니코써치", "Unico Search", "베스트커리어",
                            "비즈니스피플", "커리어앤유", "커리어투어",
                            "키스톤HR", "HR컨설팅그룹", "스탠다드", "Standard",
                            "피플게이트", "CareerNet", "커리어넷",
                            "커리어크로스", "네오피플", "포커스HR", "히든챔피언"]
    
    for kw in search_firm_keywords:
        if kw.lower() in name_clean.lower():
            return "서치펌"
    
    # 외국계 체크
    for kw in 외국계_리스트:
        if kw.lower() in name_clean.lower():
            return "외국계"
    
    # 영문명 외국계 체크
    if re.match(r'^[A-Za-z\s&\-\.]+$', name_clean) and len(name_clean) > 3:
        # Could be foreign company
        pass  # fall through to check 대기업 first
    
    # 대기업 체크
    for kw in 대기업_리스트:
        if kw in name_clean:
            return "대기업"
    
    # 중견 체크
    for kw in 중견_리스트:
        if kw in name_clean:
            return "중견"
    
    # 영문명이고 외국계 리스트에 없는 경우 → 외국계 가능성
    if re.match(r'^[A-Za-z\s&\-\.\(\)]+$', name_clean) and not any(c in '가나다라마바사아자차카타파하' for c in name_clean):
        return "외국계"
    
    return "중소"


def filter_job(title, company, location, employment, seniority_text=""):
    """필터링: True면 탈락"""
    if not title:
        return True, "제목 없음"
    
    title_lower = title.lower()
    company_lower = (company or "").lower()
    location_text = (location or "").lower()
    employment_text = (employment or "").lower()
    combined = f"{title_lower} {seniority_text.lower()} {employment_text}"
    
    # 개발자/연구원/마케터/일본어 체크
    for kw in BLACKLIST_TITLES:
        if kw.lower() in combined:
            return True, f"제외직무: {kw}"
    
    # 경력 5년+/매니저급 체크
    for kw in BLACKLIST_SENIORITY:
        if kw.lower() in combined:
            return True, f"경력과다: {kw}"
    
    # 계약직/인턴 체크
    for kw in BLACKLIST_EMPLOYMENT:
        if kw.lower() in employment_text or kw.lower() in combined:
            return True, f"고용형태: {kw}"
    
    # 비수도권 체크
    for kw in BLACKLIST_LOCATION:
        if kw.lower() in location_text:
            return True, f"비수도권: {kw}"
    
    return False, ""


def evaluate_fit(title, company, years="2년6개월"):
    """부합도 평가 (5단계)"""
    title_lower = title.lower()
    company_lower = (company or "").lower()
    
    # High match keywords
    high_match = ["사업개발", "business development", "bd", "해외영업", "global sales",
                  "b2b", "commercial", "key account", "무역", "trading",
                  "화학", "chemical", "소재", "material", "에너지", "energy",
                  "식품", "food", "바이오", "bio", "공급망", "supply chain",
                  "partnership", "alliance", "전략", "strategy"]
    
    # Count matches
    score = 0
    for kw in high_match:
        if kw in title_lower:
            score += 1
    
    # Company prestige bonus
    회사분류 = classify_company(company or "")
    if 회사분류 == "대기업":
        score += 2
    elif 회사분류 == "외국계":
        score += 2
    elif 회사분류 == "중견":
        score += 1
    
    # 경력 적합도 (2년6개월 기준)
    if "신입" in title or "new grad" in title_lower or "associate" in title_lower:
        score += 2  # 신입 가능 → 높은 부합
    elif "경력" in title and ("2년" in title or "3년" in title or "1년" in title):
        score += 1
    
    if score >= 6:
        return "⭐⭐⭐ 완벽"
    elif score >= 4:
        return "⭐⭐ 좋음"
    elif score >= 2:
        return "⭐ 보통"
    elif score >= 1:
        return "⚠️ 가능성낮음"
    else:
        return "❌ 부적합"


def normalize_date(date_str):
    """마감일 정규화"""
    if not date_str:
        return ""
    date_str = str(date_str).strip()
    # Already YYYY-MM-DD or similar
    if re.match(r'\d{4}-\d{2}-\d{2}', date_str):
        return date_str
    # D-Day format
    if 'D-' in date_str:
        return date_str
    # MM/DD format
    if '/' in date_str:
        return date_str
    return date_str


# ─── Wanted API Scraping ──────────────────────────────────────
def scrape_wanted():
    """원티드 API로 채용공고 수집"""
    all_jobs = []
    seen_ids = set()
    
    for kw in WANTED_KEYWORDS:
        print(f"  [Wanted] 검색: {kw}")
        url = f"https://www.wanted.co.kr/api/v4/jobs?country=kr&job_sort=job.latest_order&years=-1&limit=30&query={quote(kw)}"
        
        try:
            resp = requests.get(url, headers=WANTED_HEADERS, timeout=15)
            if resp.status_code != 200:
                print(f"    ⚠️ HTTP {resp.status_code}")
                continue
            
            data = resp.json()
            jobs = data.get("data", [])
            print(f"    결과: {len(jobs)}건 (전체: {data.get('total_count', 0)})")
            
            for job in jobs:
                jid = job.get("id")
                if jid in seen_ids:
                    continue
                seen_ids.add(jid)
                
                company_info = job.get("company", {})
                company_name = company_info.get("name", "")
                position = job.get("position", "")
                location = job.get("address", {}).get("location", "") if isinstance(job.get("address"), dict) else ""
                due_date = job.get("due_time", "")
                industry = company_info.get("industry_name", "")
                
                # Extract employment type
                employment = "정규"  # default
                
                # Filter
                reject, reason = filter_job(position, company_name, location, employment)
                if reject:
                    continue
                
                규모 = classify_company(company_name)
                부합도 = evaluate_fit(position, company_name)
                
                job_link = f"https://www.wanted.co.kr/wd/{jid}"
                
                all_jobs.append({
                    "source": "원티드",
                    "id": f"wanted_{jid}",
                    "규모": 규모,
                    "회사명": company_name,
                    "직무": position,
                    "구분": "경력",
                    "고용형태": "정규",
                    "지역": location or "서울",
                    "마감일": normalize_date(due_date),
                    "스코어": 0,
                    "링크": job_link,
                    "링크_url": job_link,
                    "지원상태": "",
                    "내 부합도": 부합도,
                    "산업": industry,
                })
            
            time.sleep(0.5)  # Rate limiting
            
        except Exception as e:
            print(f"    ❌ 오류: {e}")
    
    print(f"  [Wanted] 총 수집: {len(all_jobs)}건")
    return all_jobs


# ─── Peoplenjob Scraping ──────────────────────────────────────
def scrape_peoplenjob():
    """피플앤잡 채용공고 수집"""
    all_jobs = []
    seen_hashes = set()
    
    for kw in PEOPLE_KEYWORDS:
        for page in [1, 2]:
            print(f"  [Peoplenjob] 검색: {kw} (page {page})")
            url = f"https://www.peoplenjob.com/jobs?field=all&q={quote(kw)}&page={page}"
            
            try:
                resp = requests.get(url, headers=PEOPLE_HEADERS, timeout=15)
                if resp.status_code != 200:
                    print(f"    ⚠️ HTTP {resp.status_code}")
                    continue
                
                soup = BeautifulSoup(resp.text, "html.parser")
                
                # Find job cards
                cards = soup.select(".jd-card, .job-card, article, .list-item")
                
                if not cards:
                    # Try alternative selectors
                    cards = soup.select("[class*='card'], [class*='job'], [class*='item']")
                    # Fallback: try parsing all text
                    text = resp.text
                    # Look for company names pattern
                    # Try finding all links that look like job postings
                    all_links = soup.find_all("a", href=re.compile(r"/jobs/\d+|/job/|/view/"))
                    print(f"    Found {len(all_links)} job links")
                    
                    for link in all_links:
                        card_text = link.get_text(separator=" ", strip=True)
                        if len(card_text) < 20:
                            continue
                        
                        href = link.get("href", "")
                        if not href.startswith("http"):
                            href = "https://www.peoplenjob.com" + href
                        
                        # Parse text for job info
                        # Typical format: "회사명 | 직무명 | 지역 | 경력 | 마감일"
                        parts = card_text.split("|")
                        if len(parts) >= 3:
                            company_name = parts[0].strip()
                            position = parts[1].strip() if len(parts) > 1 else ""
                            location = parts[2].strip() if len(parts) > 2 else "서울"
                            employment = parts[3].strip() if len(parts) > 3 else "정규"
                            due_date = parts[4].strip() if len(parts) > 4 else ""
                        else:
                            # Try other parsing
                            company_name = parts[0].strip()
                            position = card_text
                            location = "서울"
                            employment = "정규"
                            due_date = ""
                        
                        # Dedup
                        job_hash = hashlib.md5(f"{company_name}:{position}".encode()).hexdigest()
                        if job_hash in seen_hashes:
                            continue
                        seen_hashes.add(job_hash)
                        
                        # Filter
                        reject, reason = filter_job(position, company_name, location, employment)
                        if reject:
                            continue
                        
                        규모 = classify_company(company_name)
                        부합도 = evaluate_fit(position, company_name)
                        
                        all_jobs.append({
                            "source": "피플앤잡",
                            "id": f"people_{job_hash[:8]}",
                            "규모": 규모,
                            "회사명": company_name,
                            "직무": position[:100],
                            "구분": "경력",
                            "고용형태": employment or "정규",
                            "지역": location or "서울",
                            "마감일": normalize_date(due_date),
                            "스코어": 0,
                            "링크": "🔗 지원하기",
                            "링크_url": href,
                            "지원상태": "",
                            "내 부합도": 부합도,
                            "산업": "",
                        })
                
                else:
                    for card in cards:
                        card_text = card.get_text(separator=" ", strip=True)
                        if len(card_text) < 20:
                            continue
                        
                        # Find link
                        link_tag = card.find("a", href=True)
                        href = link_tag.get("href", "") if link_tag else ""
                        if href and not href.startswith("http"):
                            href = "https://www.peoplenjob.com" + href
                        
                        # Find company name
                        company_el = card.select_one("[class*='company'], [class*='name'], strong")
                        company_name = company_el.get_text(strip=True) if company_el else ""
                        
                        # Find position
                        position_el = card.select_one("[class*='title'], [class*='position'], h2, h3")
                        position = position_el.get_text(strip=True) if position_el else card_text[:100]
                        
                        # Default location
                        location = "서울"
                        
                        # Dedup
                        job_hash = hashlib.md5(f"{company_name}:{position}".encode()).hexdigest()
                        if job_hash in seen_hashes:
                            continue
                        seen_hashes.add(job_hash)
                        
                        # Filter
                        reject, reason = filter_job(position, company_name, location, "정규")
                        if reject:
                            continue
                        
                        규모 = classify_company(company_name)
                        부합도 = evaluate_fit(position, company_name)
                        
                        all_jobs.append({
                            "source": "피플앤잡",
                            "id": f"people_{job_hash[:8]}",
                            "규모": 규모,
                            "회사명": company_name,
                            "직무": position[:100],
                            "구분": "경력",
                            "고용형태": "정규",
                            "지역": location,
                            "마감일": "",
                            "스코어": 0,
                            "링크": "🔗 지원하기",
                            "링크_url": href,
                            "지원상태": "",
                            "내 부합도": 부합도,
                            "산업": "",
                        })
                
                time.sleep(1)
                
            except Exception as e:
                print(f"    ❌ 오류: {e}")
    
    print(f"  [Peoplenjob] 총 수집: {len(all_jobs)}건")
    return all_jobs


# ─── Main ─────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("채용공고 수집 시작")
    print(f"날짜: {TODAY}")
    print("=" * 60)
    
    # 1. Scrape Wanted
    print("\n[1/3] 원티드 API 수집 중...")
    wanted_jobs = scrape_wanted()
    
    # 2. Scrape Peoplenjob
    print("\n[2/3] 피플앤잡 수집 중...")
    people_jobs = scrape_peoplenjob()
    
    # 3. Combine and deduplicate
    print("\n[3/3] 중복 제거 및 트래커 업데이트...")
    all_new_jobs = wanted_jobs + people_jobs
    
    # Additional dedup across sources
    seen = set()
    unique_jobs = []
    for job in all_new_jobs:
        key = f"{job['회사명']}:{job['직무']}"
        if key not in seen:
            seen.add(key)
            unique_jobs.append(job)
    
    # Sort by 규모 (대기업 > 외국계 > 서치펌 > 중견 > 중소)
    규모_order = {"대기업": 0, "외국계": 1, "서치펌": 2, "중견": 3, "중소": 4}
    unique_jobs.sort(key=lambda j: (규모_order.get(j["규모"], 5), -len(j.get("내 부합도", "").replace("⭐", ""))))
    
    print(f"\n📊 수집 결과:")
    print(f"  원티드: {len(wanted_jobs)}건")
    print(f"  피플앤잡: {len(people_jobs)}건")
    print(f"  중복 제거 후: {len(unique_jobs)}건")
    
    # 규모별 통계
    from collections import Counter
    규모_stats = Counter(j["규모"] for j in unique_jobs)
    for k, v in 규모_stats.items():
        print(f"  {k}: {v}건")
    
    # 4. Update Excel tracker
    print("\n📝 엑셀 트래커 업데이트 중...")
    update_tracker(unique_jobs)
    
    print(f"\n✅ 완료! 저장 경로: {NEW_TRACKER_PATH}")
    
    # Print top recommendations
    print("\n" + "=" * 60)
    print("🏆 주요 추천 공고 TOP 15")
    print("=" * 60)
    
    top_jobs = [j for j in unique_jobs if j["규모"] in ("대기업", "외국계")]
    if len(top_jobs) < 15:
        top_jobs = unique_jobs[:15]
    else:
        top_jobs = top_jobs[:15]
    
    for i, job in enumerate(top_jobs, 1):
        print(f"  {i}. [{job['규모']}] {job['회사명']} - {job['직무'][:60]}")
        print(f"     부합도: {job['내 부합도']} | {job['링크_url'][:80]}")
    
    return unique_jobs


def update_tracker(jobs):
    """기존 엑셀 트래커에 새 공고 추가"""
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb["전체 리스트"]
    
    # 기존 데이터 읽기 (회사명+직무 키)
    existing_keys = set()
    for r in range(2, ws.max_row + 1):
        company = ws.cell(row=r, column=3).value
        position = ws.cell(row=r, column=4).value
        if company and position:
            existing_keys.add(f"{str(company).strip()}:{str(position).strip()}")
    
    # Filter out jobs already in tracker
    new_jobs = []
    for job in jobs:
        key = f"{job['회사명']}:{job['직무']}"
        if key not in existing_keys:
            new_jobs.append(job)
    
    print(f"  기존 공고: {len(existing_keys)}건")
    print(f"  신규 공고: {len(new_jobs)}건")
    
    if not new_jobs:
        print("  추가할 신규 공고가 없습니다.")
        wb.save(NEW_TRACKER_PATH)
        return
    
    # Find insertion point - after last existing data row, before any summary rows
    # Group by 규모
    규모_order = {"대기업": 0, "외국계": 1, "서치펌": 2, "중견": 3, "중소": 4}
    new_jobs.sort(key=lambda j: 규모_order.get(j["규모"], 5))
    
    # Find section boundaries
    sections = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str) and v.startswith("▌"):
            sections[v] = r
    
    # Add new section header
    insert_row = ws.max_row + 1
    section_header = f"▌🆕 {TODAY_SHORT} 신규 발굴 ({len(new_jobs)}건)"
    
    # Style for section header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    ws.insert_rows(insert_row)
    ws.cell(row=insert_row, column=1).value = section_header
    ws.cell(row=insert_row, column=1).font = header_font
    ws.cell(row=insert_row, column=1).fill = header_fill
    
    # Merge section header cells for visual
    for c in range(2, 13):
        ws.cell(row=insert_row, column=c).fill = header_fill
    
    insert_row += 1
    
    # Insert new jobs
    for i, job in enumerate(new_jobs, 1):
        row = insert_row + i - 1
        
        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = job["규모"]
        ws.cell(row=row, column=3).value = job["회사명"]
        ws.cell(row=row, column=4).value = job["직무"]
        ws.cell(row=row, column=5).value = job.get("구분", "경력")
        ws.cell(row=row, column=6).value = job.get("고용형태", "정규")
        ws.cell(row=row, column=7).value = job.get("지역", "서울")
        ws.cell(row=row, column=8).value = job.get("마감일", "")
        ws.cell(row=row, column=9).value = job.get("스코어", 0)
        
        # Hyperlink for 링크
        link_cell = ws.cell(row=row, column=10)
        link_text = "🔗 지원하기"
        link_url = job.get("링크_url", "")
        if link_url:
            link_cell.value = link_text
            link_cell.hyperlink = link_url
            link_cell.font = Font(color="0563C1", underline="single")
        else:
            link_cell.value = link_text
        
        ws.cell(row=row, column=11).value = job.get("지원상태", "")
        ws.cell(row=row, column=12).value = job.get("내 부합도", "")
    
    # Also update TOP 추천 sheet
    update_top_sheet(wb, new_jobs)
    
    # Save
    wb.save(NEW_TRACKER_PATH)
    print(f"  저장 완료: {NEW_TRACKER_PATH}")


def update_top_sheet(wb, new_jobs):
    """TOP 추천 시트도 업데이트"""
    if "TOP 추천" not in wb.sheetnames:
        return
    
    ws = wb["TOP 추천"]
    
    # Find last row
    insert_row = ws.max_row + 1
    
    # Add section header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    section_header = f"▌🆕 {TODAY_SHORT} 신규 ({len(new_jobs)}건)"
    ws.insert_rows(insert_row)
    ws.cell(row=insert_row, column=1).value = section_header
    ws.cell(row=insert_row, column=1).font = header_font
    ws.cell(row=insert_row, column=1).fill = header_fill
    for c in range(2, 13):
        ws.cell(row=insert_row, column=c).fill = header_fill
    
    insert_row += 1
    
    # Only add 대기업/외국계 top jobs
    top_new = [j for j in new_jobs if j["규모"] in ("대기업", "외국계")][:10]
    
    for i, job in enumerate(top_new, 1):
        row = insert_row + i - 1
        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = job["규모"]
        ws.cell(row=row, column=3).value = job["회사명"]
        ws.cell(row=row, column=4).value = job["직무"]
        ws.cell(row=row, column=5).value = job.get("구분", "경력")
        ws.cell(row=row, column=6).value = job.get("고용형태", "정규")
        ws.cell(row=row, column=7).value = job.get("지역", "서울")
        ws.cell(row=row, column=8).value = job.get("마감일", "")
        ws.cell(row=row, column=9).value = job.get("스코어", 0)
        
        link_cell = ws.cell(row=row, column=10)
        link_url = job.get("링크_url", "")
        link_cell.value = "🔗 지원하기"
        if link_url:
            link_cell.hyperlink = link_url
            link_cell.font = Font(color="0563C1", underline="single")
        
        ws.cell(row=row, column=11).value = job.get("지원상태", "")
        ws.cell(row=row, column=12).value = job.get("내 부합도", "")


if __name__ == "__main__":
    main()
